import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

st.title("Jira LDSO Report Generator")

uploaded_file = st.file_uploader("Upload Jira Excel", type=["xlsx"])

if uploaded_file:

    df = pd.read_excel(uploaded_file)

    st.subheader("Preview Jira Data")
    st.dataframe(df)

    if st.button("Generate Report"):

        # -----------------------------
        # Sheet3 : Jira_LDSO
        # -----------------------------
        jira_ldso = df.copy()

        # -----------------------------
        # Sheet2 : Mapping
        # -----------------------------
        grouped = df.groupby("Service Name")["LDSO"].apply(list)

        rows = []
        for service, ldso_list in grouped.items():
            row = [service, len(ldso_list)] + ldso_list
            rows.append(row)

        max_len = max(len(r) for r in rows)

        for r in rows:
            r.extend([""] * (max_len - len(r)))

        mapping = pd.DataFrame(rows)
        mapping.columns = ["Service Name", "Total"] + [""]*(len(mapping.columns)-2)

        # -----------------------------
        # Summary Calculations
        # -----------------------------

        def count(type_name, rank=None, status=None):
            q = df[df["Type"] == type_name]

            if rank:
                q = q[q["Rank"] == rank]

            if status:
                q = q[q["Status"] == status]

            return len(q)

        # Incident totals
        sev_total = count("Incident","SEVERE")
        rankA_total = count("Incident","Rank A")
        rankB_total = count("Incident","Rank B")
        rankC_total = count("Incident","Rank C")

        # Closed
        sev_closed = count("Incident","SEVERE","Closed")
        rankA_closed = count("Incident","Rank A","Closed")
        rankB_closed = count("Incident","Rank B","Closed")
        rankC_closed = count("Incident","Rank C","Closed")

        # -----------------------------
        # Export Excel
        # -----------------------------
        output = BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:

            # create empty summary first
            pd.DataFrame().to_excel(writer, sheet_name="Summary", index=False)

            mapping.to_excel(writer, sheet_name="Mapping", index=False)
            jira_ldso.to_excel(writer, sheet_name="Jira_LDSO", index=False)

        wb = load_workbook(output)
        ws = wb["Summary"]

        # -----------------------------
        # Header
        # -----------------------------
        ws["A1"] = "LDCM February 2026"

        ws.append([])
        ws.append(["Incident","Total","Closed","Initial Check","Escalating","Investigating","Fixing","Under Confirmation"])

        ws.append(["SEVERE",sev_total,sev_closed])
        ws.append(["Rank A",rankA_total,rankA_closed])
        ws.append(["Rank B",rankB_total,rankB_closed])
        ws.append(["Rank C",rankC_total,rankC_closed])

        ws.append([])
        ws.append(["Change","Total","Closed","Initial Check","Escalating","Investigating","Fixing","Under Confirmation"])

        change_total = len(df[df["Type"]=="Change"])
        change_closed = len(df[(df["Type"]=="Change") & (df["Status"]=="Closed")])

        ws.append(["Request",change_total,change_closed])

        # save workbook
        final_output = BytesIO()
        wb.save(final_output)

        st.download_button(
            label="Download Report",
            data=final_output.getvalue(),
            file_name="Jira_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
